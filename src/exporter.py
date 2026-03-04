"""Export jobs to formatted Excel file."""

from __future__ import annotations

import json
import logging
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Sheet 1: Shortlist (visible jobs) ──────────────────────
SHORTLIST_COLUMNS = [
    ("S.No", 6),
    ("Status", 14),
    ("Company", 22),
    ("Job Title", 35),
    ("Experience", 28),
    ("Match Score", 12),
    ("Sponsorship", 14),
    ("Top Strengths", 40),
    ("Gaps", 40),
    ("Resume Action Plan", 50),
    ("Application Link", 50),
]

# ── Sheet 2: Filtered Out ─────────────────────────────────
FILTERED_COLUMNS = [
    ("S.No", 6),
    ("Company", 22),
    ("Job Title", 35),
    ("Source", 14),
    ("Filter Reason", 40),
    ("Required Years", 14),
    ("Application Link", 50),
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LINK_FONT = Font(color="0563C1", underline="single")

# Experience extraction patterns (reused from filters.py logic)
_YEARS_PATTERN = re.compile(r"(\d{1,2})\s*\+?\s*(?:years?|yrs?)", re.IGNORECASE)


def _parse_json_field(val: str | list) -> str:
    """Convert a JSON string or list to a comma-separated string."""
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    if isinstance(val, str):
        try:
            parsed = json.loads(val)
            if isinstance(parsed, list):
                return ", ".join(str(v) for v in parsed)
        except (json.JSONDecodeError, TypeError):
            pass
    return str(val) if val else ""


def _extract_required_years(job: dict) -> str:
    """Extract required years from title + description snippet."""
    text = (job.get("title", "") + " " + job.get("description_snippet", "")).lower()
    matches = _YEARS_PATTERN.findall(text)
    if matches:
        return f"{max(int(m) for m in matches)}+"
    return ""


def _write_header(ws, columns: list[tuple[str, int]]) -> None:
    """Write a styled header row to a worksheet."""
    for col_idx, (name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def _write_link_cell(ws, row: int, col: int, url: str) -> None:
    """Write a URL cell with hyperlink formatting."""
    cell = ws.cell(row=row, column=col, value="Apply" if url else "")
    if url:
        cell.hyperlink = url
        cell.font = LINK_FONT


def _sponsorship_label(status: str) -> str:
    """Convert sponsorship_status to display label."""
    return {"sponsored": "Yes", "not_sponsored": "No"}.get(status, "Unknown")


def _write_shortlist_row(ws, row_idx: int, sno: int, job: dict) -> None:
    """Write a single row to the Shortlist sheet."""
    col = 1
    ws.cell(row=row_idx, column=col, value=sno); col += 1
    ws.cell(row=row_idx, column=col, value=job.get("job_status", "Discovered")); col += 1
    ws.cell(row=row_idx, column=col, value=job.get("company", "")); col += 1
    ws.cell(row=row_idx, column=col, value=job.get("title", "")); col += 1
    ws.cell(row=row_idx, column=col, value=job.get("experience_alignment", "")); col += 1
    ws.cell(row=row_idx, column=col, value=job.get("match_score", 0)); col += 1
    ws.cell(row=row_idx, column=col, value=_sponsorship_label(job.get("sponsorship_status", "unknown"))); col += 1
    ws.cell(row=row_idx, column=col, value=_parse_json_field(job.get("strong_matches", "[]"))); col += 1
    ws.cell(row=row_idx, column=col, value=_parse_json_field(job.get("missing_keywords", "[]"))); col += 1
    ws.cell(row=row_idx, column=col, value=job.get("resume_improvement_prompt", "")); col += 1
    _write_link_cell(ws, row_idx, col, job.get("url", ""))


def _write_filtered_row(ws, row_idx: int, sno: int, job: dict) -> None:
    """Write a single row to the Filtered Out sheet."""
    col = 1
    ws.cell(row=row_idx, column=col, value=sno); col += 1
    ws.cell(row=row_idx, column=col, value=job.get("company", "")); col += 1
    ws.cell(row=row_idx, column=col, value=job.get("title", "")); col += 1
    ws.cell(row=row_idx, column=col, value=job.get("source", "")); col += 1
    ws.cell(row=row_idx, column=col, value=job.get("filter_reason", "")); col += 1
    ws.cell(row=row_idx, column=col, value=_extract_required_years(job)); col += 1
    _write_link_cell(ws, row_idx, col, job.get("url", ""))


def export_to_excel(jobs: list[dict], output_path: Path) -> Path:
    """Create a formatted Excel workbook from job data.

    Sheet 1 ("Shortlist"): only visible jobs — clean columns.
    Sheet 2 ("Filtered Out"): hidden jobs with their filter reason.

    Args:
        jobs: list of job dicts (from DB rows).
        output_path: where to write the .xlsx file.

    Returns:
        The output path.
    """
    # Split into visible vs filtered
    visible = []
    filtered = []
    for j in jobs:
        is_vis = j.get("is_visible", True)
        if is_vis in (True, 1):
            visible.append(j)
        else:
            filtered.append(j)

    wb = Workbook()

    # ── Sheet 1: Shortlist (visible jobs) ──────────────────
    ws = wb.active
    ws.title = "Shortlist"
    _write_header(ws, SHORTLIST_COLUMNS)

    for idx, job in enumerate(visible, start=1):
        _write_shortlist_row(ws, idx + 1, idx, job)

    last_col = get_column_letter(len(SHORTLIST_COLUMNS))
    last_row = max(ws.max_row, 2)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # ── Sheet 2: Filtered Out ──────────────────────────────
    ws2 = wb.create_sheet("Filtered Out")
    _write_header(ws2, FILTERED_COLUMNS)

    for idx, job in enumerate(filtered, start=1):
        _write_filtered_row(ws2, idx + 1, idx, job)

    last_col2 = get_column_letter(len(FILTERED_COLUMNS))
    last_row2 = max(ws2.max_row, 2)
    ws2.auto_filter.ref = f"A1:{last_col2}{last_row2}"

    wb.save(str(output_path))
    logger.info(
        f"Excel exported: {output_path} "
        f"({len(visible)} shortlisted, {len(filtered)} filtered)"
    )
    return output_path
